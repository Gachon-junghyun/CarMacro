# -*- coding: utf-8 -*-
"""세로형(한 파일=한 신청서) 양식 파서.

라벨 셀을 찾아 같은 행 오른쪽의 값을 읽는다. 안전하게 매핑되는 값만
site 필드 id 로 변환하고, 추측이 위험한 항목은 handoff(사람이 직접) 로 분리한다.
"""
import re
import openpyxl


def _norm(s):
    return re.sub(r"\s+", "", str(s or "")).strip()


def _is_empty_val(v):
    s = str(v or "").strip()
    return s == "" or s == "."


def _date_norm(v):
    """2026.6.17 / 2026-6-17 / 2026년 6월 17일 / 20260617 → 2026-06-17"""
    s = str(v or "").strip()
    parts = [p for p in re.split(r"\D+", s) if p]
    if len(parts) == 3:
        y, m, d = parts
        if len(y) == 4:
            return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
    digits = re.sub(r"\D", "", s)
    if len(digits) == 8:
        return f"{digits[0:4]}-{digits[4:6]}-{digits[6:8]}"
    return s


def _birth6(v):
    """1978-06-11 → 780611 (YYMMDD)"""
    digits = re.sub(r"\D", "", str(v or ""))
    if len(digits) == 8:       # YYYYMMDD
        return digits[2:8]
    if len(digits) == 6:
        return digits
    return ""


# 값 변환 사전 (정규화된 한글 → site 코드)
MOBILE_RE = re.compile(r"01[016789]-?\d{3,4}-?\d{4}")
REQ_KIND = {"개인": "P", "개인사업자": "B", "단체": "G"}
# 단체 신청구분(grp_reqst_se): 공공기관 01 / 지자체 02 / 기타 99
GRP_REQST = {"공공기관": "01", "지자체": "02", "기타": "99"}
SEX = {"남": "M", "남자": "M", "여": "F", "여자": "F"}
PRIORITY = {"우선": "10", "우선순위": "10", "법인기관": "20", "법인·기관": "20",
            "중소기업": "30", "택배": "50", "일반": "00"}
YN = {"Y": "Y", "N": "N", "예": "Y", "아니오": "N", "O": "Y"}
# 유종 한글 → 라이브 폼 fuel select 코드 (LP=LPG, DS=경유, GS=휘발유)
FUEL_MAP = {"LPG": "LP", "엘피지": "LP", "액화석유가스": "LP", "LPG(액화석유가스)": "LP",
            "경유": "DS", "디젤": "DS", "휘발유": "GS", "가솔린": "GS"}


def parse_vertical(path):
    """반환: dict(mapped, handoff, notes, raw)
    - mapped:  site 필드 id -> 값 (자동 입력 대상)
    - special: 텍스트 매칭이 필요한 항목 (model_cd, social_kind 한글명)
    - handoff: [(라벨, 값, 사유)] 사람이 직접 확인/입력
    - notes:   파싱 경고
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # 셀 그리드: (row, col) -> 문자열
    grid = {}
    maxr = ws.max_row
    maxc = ws.max_column
    for r in range(1, maxr + 1):
        for c in range(1, maxc + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                grid[(r, c)] = str(v).strip()

    def value_right(r, c):
        """같은 행에서 (r,c) 오른쪽 첫 비어있지 않은 셀 값."""
        for cc in range(c + 1, maxc + 1):
            if (r, cc) in grid:
                return grid[(r, cc)], cc
        return "", None

    # 라벨 위치 모으기: 정규화 라벨 -> [(r,c), ...]
    label_pos = {}
    for (r, c), v in grid.items():
        label_pos.setdefault(_norm(v), []).append((r, c))

    mapped, special, handoff, notes = {}, {}, [], []
    used = set()  # 값으로 소비한 셀 좌표 (중복 방지)

    def get(label):
        """라벨에 매칭되는 첫 값과 좌표. 없으면 ('', None)."""
        for key, positions in label_pos.items():
            if key == _norm(label):
                for (r, c) in positions:
                    val, vc = value_right(r, c)
                    if vc is not None:
                        return val, (r, vc)
        return "", None

    def get_all(label):
        """동일 라벨이 여러 번일 때 모든 (값, 좌표)."""
        res = []
        for (r, c) in label_pos.get(_norm(label), []):
            val, vc = value_right(r, c)
            if vc is not None:
                res.append((val, (r, vc)))
        return res

    def mobile_in_row_of(label):
        """라벨이 있는 행에서 휴대폰 패턴 값을 찾음(라벨 없는 신청자 휴대폰용)."""
        for (r, c) in label_pos.get(_norm(label), []):
            for cc in range(1, maxc + 1):
                val = grid.get((r, cc), "")
                m = MOBILE_RE.search(val.replace(" ", ""))
                if m:
                    return m.group(0)
        return ""

    # ── 안전 매핑 (텍스트/확정 코드) ──────────────────────────
    # 전기/수소 구분 (열린 폼과 대조용 — 사고 방지)
    v, _ = get("구분")
    if v:
        nv = _norm(v)
        if "수소" in nv:
            mapped["car_kind"] = "수소"
        elif "전기" in nv:
            mapped["car_kind"] = "전기"

    v, _ = get("계약일자")
    if v:
        mapped["contract_day"] = _date_norm(v)

    v, _ = get("신청유형")
    if v:
        code = REQ_KIND.get(_norm(v))
        if code:
            mapped["req_kind"] = code
        else:
            handoff.append(("신청유형", v, "req_kind 코드 매칭 실패"))

    v, _ = get("성명")
    if v:
        mapped["req_nm"] = v

    v, _ = get("생년월일")
    if v:
        d8 = _date_norm(v)   # 1978-06-11 / 1982.3.22 → YYYY-MM-DD
        if re.match(r"^\d{4}-\d{2}-\d{2}$", d8):
            mapped["birth1"] = d8
        elif _birth6(v):
            mapped["birth1"] = _birth6(v)   # 6자리만 있으면 그대로
        else:
            handoff.append(("생년월일", v, "형식 불명 → 직접 확인"))

    v, _ = get("성별")
    if v:
        code = SEX.get(_norm(v))
        if code:
            mapped["req_sex"] = code
        else:
            handoff.append(("성별", v, "M/F 매칭 실패"))

    # 사업자 정보
    v, _ = get("사업자번호")
    if v:
        mapped["busi_no"] = v
    v, _ = get("사업자명")
    if v:
        mapped["pri_busi_nm"] = v

    # ── 단체(법인·기관) 전용 라벨 → 폼 필드 id 매핑 ──
    #  단체 폼은 같은 칸 id 가 개인과 다른 뜻을 가진다:
    #    req_nm=기관명(개인은 성명) · ceo=대표자 · birth2=법인등록번호(개인은 주민번호)
    v, _ = get("기관명")
    if v:
        mapped["req_nm"] = v            # 단체: 기관명이 성명(req_nm) 칸에 들어감
    v, _ = get("대표자")
    if v:
        mapped["ceo"] = v
    v, _ = get("법인등록번호")
    if v:
        mapped["birth2"] = v            # 단체: 법인등록번호는 birth2 칸
    v, _ = get("신청구분")
    if v:
        code = GRP_REQST.get(_norm(v))
        if code:
            mapped["grp_reqst_se"] = code
        else:
            handoff.append(("신청구분", v, "grp_reqst_se 매칭 실패"))
    v, _ = get("개인사업장명")   # 단체 폼의 사업자명 라벨
    if v:
        mapped["pri_busi_nm"] = v

    # 신청차종 → 라이브 옵션 텍스트 매칭 필요
    v, _ = get("신청차종")
    if v:
        special["model_cd"] = v   # 한글 차종명

    v, _ = get("신청대수")
    if v:
        mapped["req_cnt"] = re.sub(r"\D", "", v) or v

    v, _ = get("출고예정일자")
    if v:
        mapped["delivery_sch_day"] = _date_norm(v)

    # 주소
    addr_val, addr_pos = get("주소")
    if addr_val:
        mapped["addr"] = addr_val
    v, _ = get("이하주소")
    if v:
        mapped["addr_detail"] = v
    elif addr_pos:
        # 라벨 없는 양식: 주소(도로명) 값 오른쪽 다음 셀을 상세주소로 사용
        ar, ac = addr_pos
        for cc in range(ac + 1, maxc + 1):
            if (ar, cc) in grid:
                mapped["addr_detail"] = grid[(ar, cc)]
                break
    # 우편번호는 데이터에 없음 → 주소(도로명)로 검색 팝업 자동화(결과 1개일 때 자동 선택)
    notes.append("※ 우편번호는 '주소'(도로명)로 주소검색 팝업을 자동 실행합니다. "
                 "결과가 정확히 1개면 자동 선택, 여러 개/0개면 팝업에서 직접 선택.")

    # 연락처: 휴대폰은 '신청자 휴대폰'(mobile), 전화번호 '.'은 빈값
    # 신청자 휴대폰 = E11 의 010-... (라벨 없이 값만). 담당자 휴대폰과 구분 주의.
    # 전화/이메일은 필수(*) 칸이라 데이터의 '.' 도 그대로 입력(빈칸만 건너뜀)
    v, _ = get("전화번호")
    if not str(v).strip():
        v, _ = get("전화")
    if str(v).strip():
        mapped["phone"] = str(v).strip()
    v, _ = get("이메일")
    if str(v).strip():
        mapped["email"] = str(v).strip()
    # 신청자 휴대폰:
    #  - (구양식) 전화/이메일 행에 라벨 없이 들어있는 010-... 값
    #  - (실양식) '휴대폰' 라벨이 신청자용 (별도로 '연락담당휴대폰' 라벨이 존재)
    mob = (mobile_in_row_of("이메일") or mobile_in_row_of("전화번호")
           or mobile_in_row_of("전화"))
    if not mob and label_pos.get(_norm("연락담당휴대폰")):
        hv, _ = get("휴대폰")
        m = MOBILE_RE.search(str(hv).replace(" ", ""))
        if m:
            mob = m.group(0)
    if mob:
        mapped["mobile"] = mob

    # 사회계층여부 (라벨 2개: 여부 Y/N + 유형 한글)
    socs = get_all("사회계층여부")
    social_yn = ""
    social_kind_txt = ""
    for val, _pos in socs:
        nv = _norm(val)
        if nv in YN:
            social_yn = YN[nv]
        elif val.strip():
            social_kind_txt = val.strip()
    if social_yn:
        mapped["social_yn"] = social_yn
    if social_yn == "Y" and social_kind_txt:
        special["social_kind"] = social_kind_txt   # '소상공인' 등 텍스트 매칭

    # 우선순위
    v, _ = get("우선순위 배정.집행 선택")
    if not v:
        v, _ = get("우선순위")
    if v:
        code = PRIORITY.get(_norm(v))
        if code:
            mapped["priority_type"] = code
        else:
            handoff.append(("우선순위", v, "priority_type 매칭 실패"))

    # 연락담당 성명 → contact_nm (구양식: '담당자성명')
    v, _ = get("연락담당성명")
    if not v:
        v, _ = get("담당자성명")
    if v:
        mapped["contact_nm"] = v
    # 연락담당 휴대폰 → contact_mobile
    #  - 실양식: '연락담당휴대폰' 라벨
    #  - 구양식: '휴대폰' 라벨 (이때 신청자 휴대폰은 라벨 없는 값에서 읽음)
    v, _ = get("연락담당휴대폰")
    if not v:
        v, _ = get("휴대폰")
    if v and MOBILE_RE.search(v.replace(" ", "")):
        mapped["contact_mobile"] = v
    # 제조수입사 관리(계약)번호 → seller_mgrid (구양식: '계약번호')
    v, _ = get("제조수입사계약번호")
    if not v:
        v, _ = get("계약번호")
    if v:
        mapped["seller_mgrid"] = v

    # 택시 여부 (개인택시 등)
    v, _ = get("택시여부")
    if v and _norm(v) in YN:
        mapped["taxi_yn"] = YN[_norm(v)]

    # ── 전기 화물(car_type=12) 전용 항목 ──
    # 농업인 여부
    v, _ = get("농업인여부")
    if v and _norm(v) in YN:
        mapped["farmng_yn"] = YN[_norm(v)]
    # 택배여부
    v, _ = get("택배여부")
    if v and _norm(v) in YN:
        mapped["hdry_yn"] = YN[_norm(v)]

    # 생애최초 차량 구매자 여부
    v, _ = get("생애최초 차량 구매자여부")
    if not v:
        v, _ = get("생애최초")
    if v and _norm(v) in YN:
        mapped["first_buy_yn"] = YN[_norm(v)]

    # ── 전환지원금(노후 내연기관차 폐차 후 전기차 구매) + 폐차 정보 ──
    #  라이브 폼: exchange_3year_yn=Y → 폐차대수 입력 → create3yearNewCarInfo() 로
    #  '전환지원금 폐차 정보' 행이 생성됨. 행 입력칸은 id 없이 name 기반:
    #  bf_owner_nm/exchg_delivery_day/own_start_dt/own_end_dt/fuel/exchg_vh_num
    ex_consumed_carno = False
    v, _ = get("전환지원금")
    if v and _norm(v) in YN:
        ex = YN[_norm(v)]
        mapped["exchange_3year_yn"] = ex
        if ex == "Y":
            mapped["exchange_3year_cnt"] = "1"   # 폐차대수(기본 1대)
            scrap = {}
            po, _ = get("직전소유주")
            if po:
                scrap["bf_owner_nm"] = po
            fr, _ = get("최초등록일")
            if not fr:
                fr, _ = get("차량최초등록일")
            if fr:
                scrap["exchg_delivery_day"] = _date_norm(fr)
            os_, _ = get("소유기간시작일")
            if os_:
                scrap["own_start_dt"] = _date_norm(os_)
            oe, _ = get("소유기간종료일")
            if oe:
                scrap["own_end_dt"] = _date_norm(oe)
            fu, _ = get("유종")
            if fu:
                scrap["fuel"] = FUEL_MAP.get(_norm(fu), fu.strip())
            cn, _ = get("차량번호")
            if cn:
                scrap["exchg_vh_num"] = cn.strip()
                ex_consumed_carno = True   # 전환지원금 폐차차량번호로 소비 → handoff 안 함
            if scrap:
                mapped["exchange_scrap"] = scrap

    # ── 경유화물차 보유 미이행자 정보 (thrgh_ex_yn=Y → 보유차량 동적행) ──
    #   라이브 폼: thrgh_ex_yn=Y → 보유대수(thrgh_ex_cnt) → createNewCarInfoHold() 로 행 생성.
    #   생성칸(name): ex_vh_num_hold(차량번호)/ex_vh_id_hold(차대번호)/ex_model_nm_hold(보유모델명)
    v, _ = get("경유화물차 폐차 미이행여부")
    if not v:
        v, _ = get("경유화물차 미이행자 여부")
    if v and _norm(v) in YN:
        thy = YN[_norm(v)]
        mapped["thrgh_ex_yn"] = thy
        if thy == "Y":
            mapped["thrgh_ex_cnt"] = "1"   # 보유대수(기본 1대)
            hold = {}
            cn, _ = get("차량번호")   # 보유차량번호(전환지원금 폐차차량번호로 이미 소비됐으면 제외)
            if cn and not ex_consumed_carno:
                hold["ex_vh_num_hold"] = cn.strip()
            vin, _ = get("차대번호")
            if not vin:
                vin, _ = get("보유차대번호")
            if vin:
                hold["ex_vh_id_hold"] = vin.strip()
            hm, _ = get("보유모델명")
            if hm:
                hold["ex_model_nm_hold"] = hm.strip()
            if hold:
                mapped["thrgh_hold"] = hold

    # 암호 역순
    handoff.append(("암호 역순 입력", "제출 직전 단계", "캡차성 단계 → 사람이 직접 (자동화 안 함)"))

    # '구분' 칸이 없을 때: 수소 전용 차종명(넥쏘 등)이면 수소로 추정(안전가드용)
    if "car_kind" not in mapped:
        mtxt = _norm(special.get("model_cd", ""))
        if "넥쏘" in mtxt or "수소" in mtxt:
            mapped["car_kind"] = "수소"
            notes.append("※ '구분' 칸이 없어 차종으로 수소차로 추정했습니다. "
                         "열린 폼이 수소(H2)인지 꼭 확인하세요.")

    return {"mapped": mapped, "special": special, "handoff": handoff, "notes": notes}


if __name__ == "__main__":
    import os
    import sys
    default = os.path.join(os.path.dirname(__file__), "examples", "real_sample_ev.xlsx")
    path = sys.argv[1] if len(sys.argv) > 1 else default
    res = parse_vertical(path)
    print("=== 자동 입력(mapped) ===")
    for k, v in res["mapped"].items():
        print(f"  {k:18} = {v}")
    print("\n=== 텍스트 매칭 필요(special) ===")
    for k, v in res["special"].items():
        print(f"  {k:18} = {v}")
    print("\n=== 사람이 직접(handoff) ===")
    for lbl, val, reason in res["handoff"]:
        print(f"  · {lbl} = '{val}'  ({reason})")
    print("\n=== 메모 ===")
    for n in res["notes"]:
        print("  ", n)
