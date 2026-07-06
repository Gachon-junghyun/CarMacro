# -*- coding: utf-8 -*-
"""신청서 이미지에서 뽑은 값(JSON 스펙) → 실전 양식 그리드 xlsx 생성기.

examples/real_sample_ev.xlsx 와 동일한 2D 그리드 배치로 저장하며,
'구분'(전기/수소) 마커를 최상단에 넣어 parser.py 가 car_kind 를 자동 인식하게 한다.

사용:
  단건:  python .claude/skills/real-dataset/scripts/build_real.py spec.json
  다건:  spec.json 에 JSON 배열([...]) 을 넣으면 각각 생성 + 통합목록도 생성
  옵션:  --outdir DIR   출력 폴더 강제 지정(기본: real_<MMDD>, 계약일자에서 유추)
         --combined     다건일 때 통합목록 강제 생성(기본: 2건 이상이면 자동)

스펙 필드(모두 문자열, 없으면 생략 가능):
  gubun          "전기" | "수소"  (생략 시 model 로 자동 추정: 넥쏘→수소, 그 외→전기)
  contract_day   계약일자         req_kind  "개인"|"개인사업자"|"단체"
  name 성명       birth 생년월일    sex 성별("남"/"여")
  busi_no 사업자번호   busi_nm 사업자명   (개인사업자일 때)
  model 신청차종  cnt 신청대수   delivery 출고예정일자   subsidy 보조금금액
  addr 주소       addr_detail 이하주소(상세)
  phone 전화번호  mobile 신청자휴대폰(라벨없이 입력)  email 이메일
  taxi_yn 택시여부   first_buy_yn 생애최초차량구매자여부   social_yn 사회계층여부   social_kind 사회계층유형
  exchange_yn 전환지원금(Y/N)  prev_owner 직전소유주  first_reg 차량최초등록일
    own_start 소유기간시작일  own_end 소유기간종료일  fuel 유종(LPG/경유/휘발유)
    scrap_car_no (전환지원금)폐차 차량번호
  scrap_yn 경유화물차폐차미이행여부  car_no 차량번호  vin 차대번호  hold_model 보유모델명
  priority 우선순위   mfr_contact 제조사연락처
  contact_nm 담당자성명   contact_mobile 담당자휴대폰   contract_no 계약번호
  pw 암호   pw_rev 암호역순
"""
import os
import re
import sys
import json
import openpyxl
from openpyxl.styles import Font

# 수소차(FCEV) 모델 키워드 — 그 외에는 전기(EV)로 분류
H2_KEYWORDS = ("넥쏘", "디올넥쏘", "수소", "fcev", "h2")


def derive_gubun(spec):
    g = str(spec.get("gubun") or "").strip()
    if g:
        return "수소" if "수소" in g else ("전기" if "전기" in g else g)
    model = re.sub(r"\s+", "", str(spec.get("model") or "")).lower()
    for kw in H2_KEYWORDS:
        if kw in model:
            return "수소"
    return "전기"  # 기본값: 전기차


def _mmdd(spec):
    d = str(spec.get("contract_day") or spec.get("delivery") or "")
    digits = re.sub(r"\D", "", d)
    return digits[4:8] if len(digits) >= 8 else "0000"


def build_workbook(spec):
    """스펙 dict → openpyxl Workbook (실전 그리드 배치)."""
    gubun = derive_gubun(spec)
    g = spec.get  # 짧은 별칭

    # 각 줄(line)은 최대 3개 슬롯: slot0→B/C, slot1→D/E, slot2→F/G
    # 슬롯 = (라벨, 값). None 슬롯은 건너뜀. 라벨만/값만도 허용.
    lines = [
        [("구분", gubun)],
        [("계약일자", g("contract_day"))],
        [("신청유형", g("req_kind"))],
        [("성명", g("name")), ("생년월일", g("birth")), ("성별", g("sex"))],
    ]
    if g("req_kind") == "개인사업자" or g("busi_no") or g("busi_nm"):
        lines.append([("개인사업자", None), ("사업자번호", g("busi_no")),
                      ("사업자명", g("busi_nm"))])
    lines.append([("신청차종", g("model")), ("신청대수", g("cnt")),
                  ("출고예정일자", g("delivery"))])
    if g("subsidy"):
        lines.append([None, None, ("보조금금액", g("subsidy"))])
    lines.append([("주소", g("addr"))])
    lines.append([("이하주소", g("addr_detail"))])
    # 전화번호 / (라벨 없는 신청자 휴대폰) / 이메일 — parser 는 이 행에서 휴대폰 패턴을 찾음
    lines.append([("전화번호", g("phone")), (None, g("mobile")), ("이메일", g("email"))])
    if g("first_buy_yn"):
        lines.append([("생애최초 차량 구매자여부", g("first_buy_yn"))])
    if g("taxi_yn") or g("social_yn") or g("social_kind"):
        lines.append([("택시여부", g("taxi_yn")), ("사회계층여부", g("social_yn")),
                      (None, g("social_kind"))])
    # 전환지원금(노후 내연기관차 폐차 후 전기차 구매) — 폐차 정보 포함
    if g("exchange_yn"):
        lines.append([("전환지원금", g("exchange_yn")), ("사회계층여부", g("social_yn"))])
        lines.append([("직전소유주", g("prev_owner")), ("최초등록일", g("first_reg")),
                      ("소유기간시작일", g("own_start"))])
        lines.append([("소유기간종료일", g("own_end")), ("유종", g("fuel")),
                      ("차량번호", g("scrap_car_no"))])
    if g("scrap_yn"):
        lines.append([("경유화물차 폐차 미이행여부", g("scrap_yn"))])
        lines.append([("차량번호", g("car_no")), ("차대번호", g("vin")),
                      ("보유모델명", g("hold_model"))])
    if g("priority"):
        lines.append([("우선순위 배정.집행 선택", g("priority"))])
    lines.append([("제조사연락처", None), None,
                  ("전화번호", g("mfr_contact") or "공란 가능")])
    lines.append([("담당자성명", g("contact_nm")), ("휴대폰", g("contact_mobile")),
                  ("계약번호", g("contract_no"))])
    if g("pw") or g("pw_rev"):
        lines.append([(g("pw"), g("pw_rev"))])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "신청서"

    def nonempty(x):
        return x not in (None, "")

    r = 2
    for line in lines:
        has = any(s and (nonempty(s[0]) or nonempty(s[1])) for s in line)
        if not has:
            continue
        for i, slot in enumerate(line):
            if not slot:
                continue
            lbl, val = slot
            if nonempty(lbl):
                cell = ws.cell(row=r, column=2 + i * 2, value=lbl)
                if lbl in ("신청유형",):
                    cell.font = Font(bold=True)
            if nonempty(val):
                ws.cell(row=r, column=3 + i * 2, value=val)
        r += 1

    for col, w in {"B": 16, "C": 40, "D": 12, "E": 20, "F": 12, "G": 16}.items():
        ws.column_dimensions[col].width = w
    return wb


# ── 통합목록(신청서목록) : parser 로 개별파일을 재파싱해 코드값으로 변환 ──
COMBINED_HEAD = ["car_kind", "contract_day", "req_kind", "req_nm", "birth1",
                 "req_sex", "req_cnt", "delivery_sch_day", "addr", "addr_detail",
                 "phone", "email", "mobile", "contact_nm", "contact_mobile",
                 "seller_mgrid", "model_cd"]


def _find_project_root(start):
    """parser.py 가 있는 상위 폴더를 찾아 반환(없으면 cwd)."""
    d = os.path.abspath(start)
    for _ in range(8):
        if os.path.exists(os.path.join(d, "parser.py")):
            return d
        nd = os.path.dirname(d)
        if nd == d:
            break
        d = nd
    return os.getcwd()


def build_combined(paths, out_path):
    root = _find_project_root(os.getcwd())
    if root not in sys.path:
        sys.path.insert(0, root)
    from parser import parse_vertical  # noqa: E402

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "신청서목록"
    ws.append(COMBINED_HEAD)
    for p in paths:
        res = parse_vertical(p)
        m, s = res["mapped"], res["special"]
        row = [m.get("car_kind", ""), m.get("contract_day", ""), m.get("req_kind", ""),
               m.get("req_nm", ""), m.get("birth1", ""), m.get("req_sex", ""),
               m.get("req_cnt", ""), m.get("delivery_sch_day", ""), m.get("addr", ""),
               m.get("addr_detail", ""), m.get("phone", ""), m.get("email", ""),
               m.get("mobile", ""), m.get("contact_nm", ""), m.get("contact_mobile", ""),
               m.get("seller_mgrid", ""), s.get("model_cd", "")]
        ws.append(row)
    wb.save(out_path)
    return out_path


def main(argv):
    args = [a for a in argv if not a.startswith("--")]
    opts = [a for a in argv if a.startswith("--")]
    outdir_override = None
    force_combined = "--combined" in opts
    for o in opts:
        if o.startswith("--outdir"):
            # --outdir=DIR 또는 --outdir DIR 둘 다 지원
            if "=" in o:
                outdir_override = o.split("=", 1)[1]
            else:
                idx = argv.index(o)
                if idx + 1 < len(argv):
                    outdir_override = argv[idx + 1]
                    if outdir_override in args:
                        args.remove(outdir_override)

    if not args:
        print("사용법: build_real.py spec.json [--outdir DIR] [--combined]")
        return 1
    spec_path = args[0]
    with open(spec_path, encoding="utf-8") as f:
        data = json.load(f)

    people = data if isinstance(data, list) else [data]
    base = os.getcwd()
    created = []
    for spec in people:
        outdir = outdir_override or os.path.join(base, f"real_{_mmdd(spec)}")
        os.makedirs(outdir, exist_ok=True)
        name = str(spec.get("name") or "무명").strip()
        gubun = derive_gubun(spec)
        wb = build_workbook(spec)
        path = os.path.join(outdir, f"신청서_{name}.xlsx")
        wb.save(path)
        created.append(path)
        print(f"[{gubun}] created: {path}")

    if created and (force_combined or len(created) > 1):
        outdir = os.path.dirname(created[0])
        cpath = os.path.join(outdir, f"신청서_통합_{len(created)}건.xlsx")
        build_combined(created, cpath)
        print(f"[통합] created: {cpath}")

    print(f"\n총 {len(created)}건 생성 완료.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
